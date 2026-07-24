from __future__ import annotations

import base64
import hashlib
import hmac
import html
import re
import threading
from io import BytesIO
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import get_settings


SEARCH_AD_URL = (
    "https://api.searchad.naver.com/keywordstool"
)
NAVER_SHOPPING_URL = (
    "https://openapi.naver.com/v1/search/shop.json"
)

CATEGORY_CACHE_SECONDS = 60 * 60 * 6
_category_cache: dict[
    str,
    tuple[float, dict[str, Any]],
] = {}
_category_cache_lock = threading.Lock()

SHOPPING_REQUEST_INTERVAL = 0.2
_shopping_request_lock = threading.Lock()
_last_shopping_request_at = 0.0


class KeywordAnalysisError(Exception):
    pass


def normalize_keyword(value: Any) -> str:
    return re.sub(
        r"\s+",
        "",
        str(value or "").strip(),
    )


def parse_volume(value: Any) -> tuple[int, bool]:
    text = str(value or "").strip()

    if text.replace(" ", "") in {
        "<10",
        "10미만",
    }:
        return 5, True

    try:
        return (
            int(
                float(
                    text.replace(",", "")
                )
            ),
            False,
        )
    except (TypeError, ValueError):
        return 0, False


def make_search_ad_signature(
    timestamp: str,
    method: str,
    uri: str,
    secret_key: str,
) -> str:
    message = (
        f"{timestamp}.{method}.{uri}"
    )

    return base64.b64encode(
        hmac.new(
            secret_key.encode("utf-8"),
            message.encode("utf-8"),
            hashlib.sha256,
        ).digest()
    ).decode("ascii")


def fetch_keyword_statistics(
    keyword: str,
) -> list[dict[str, Any]]:
    settings = get_settings()

    if not settings.keyword_api_settings_ready:
        raise KeywordAnalysisError(
            "네이버 검색광고 API 설정이 필요합니다."
        )

    timestamp = str(
        int(time.time() * 1000)
    )
    uri = "/keywordstool"

    signature = make_search_ad_signature(
        timestamp=timestamp,
        method="GET",
        uri=uri,
        secret_key=settings.naver_ad_secret_key,
    )

    try:
        response = requests.get(
            SEARCH_AD_URL,
            headers={
                "X-Timestamp": timestamp,
                "X-API-KEY": (
                    settings.naver_ad_access_license
                ),
                "X-Customer": (
                    settings.naver_ad_customer_id
                ),
                "X-Signature": signature,
                "Content-Type": "application/json",
            },
            params={
                "hintKeywords": (
                    normalize_keyword(keyword)
                ),
                "showDetail": "1",
            },
            timeout=(5, 30),
        )
    except requests.RequestException as error:
        raise KeywordAnalysisError(
            "검색광고 API에 연결하지 못했습니다."
        ) from error

    if response.status_code != 200:
        try:
            data = response.json()
            message = (
                data.get("title")
                or data.get("detail")
                or str(data)
            )
        except Exception:
            message = response.text[:300]

        raise KeywordAnalysisError(
            f"검색광고 API 오류 "
            f"{response.status_code}: {message}"
        )

    data = response.json()
    keyword_list = data.get(
        "keywordList",
        [],
    )

    if not isinstance(keyword_list, list):
        raise KeywordAnalysisError(
            "검색광고 API 응답 형식이 올바르지 않습니다."
        )

    return keyword_list


def clean_category(value: Any) -> str:
    return html.unescape(
        re.sub(
            r"</?b>",
            "",
            str(value or ""),
            flags=re.IGNORECASE,
        )
    ).strip()


def wait_for_shopping_request_slot() -> None:
    global _last_shopping_request_at

    with _shopping_request_lock:
        current_time = time.monotonic()
        elapsed = (
            current_time
            - _last_shopping_request_at
        )
        wait_seconds = max(
            0.0,
            SHOPPING_REQUEST_INTERVAL - elapsed,
        )

        if wait_seconds:
            time.sleep(wait_seconds)

        _last_shopping_request_at = (
            time.monotonic()
        )


def fetch_shopping_category(
    keyword: str,
) -> dict[str, Any]:
    settings = get_settings()

    if (
        not settings.naver_client_id
        or not settings.naver_client_secret
    ):
        raise KeywordAnalysisError(
            "네이버 쇼핑 API 설정이 필요합니다."
        )

    response = None
    last_error: Exception | None = None

    for attempt in range(5):
        wait_for_shopping_request_slot()

        try:
            response = requests.get(
                NAVER_SHOPPING_URL,
                headers={
                    "X-Naver-Client-Id": (
                        settings.naver_client_id
                    ),
                    "X-Naver-Client-Secret": (
                        settings.naver_client_secret
                    ),
                    "User-Agent": (
                        "Fishingtem-Keyword-Analysis/2.0"
                    ),
                },
                params={
                    "query": keyword,
                    "display": 100,
                    "start": 1,
                    "sort": "sim",
                },
                timeout=(5, 25),
            )
        except requests.RequestException as error:
            last_error = error

            if attempt < 4:
                time.sleep(1 + attempt)
                continue

            raise KeywordAnalysisError(
                "네이버 쇼핑 API에 연결하지 못했습니다."
            ) from error

        if response.status_code == 200:
            break

        if response.status_code == 429:
            retry_after_text = (
                response.headers.get(
                    "Retry-After",
                    "",
                )
            )

            try:
                retry_after = float(
                    retry_after_text
                )
            except (TypeError, ValueError):
                retry_after = min(
                    8.0,
                    1.5 * (attempt + 1),
                )

            if attempt < 4:
                time.sleep(
                    max(1.0, retry_after)
                )
                continue

        raise KeywordAnalysisError(
            f"네이버 쇼핑 API 오류 "
            f"{response.status_code}"
        )

    if response is None:
        raise KeywordAnalysisError(
            "네이버 쇼핑 API 응답이 없습니다."
        ) from last_error

    if response.status_code != 200:
        raise KeywordAnalysisError(
            f"네이버 쇼핑 API 오류 "
            f"{response.status_code}"
        )

    data = response.json()
    items = data.get("items", [])
    category_counter: Counter[str] = Counter()

    for item in items:
        categories = [
            clean_category(
                item.get(f"category{number}")
            )
            for number in range(1, 5)
        ]

        category_path = " > ".join(
            category
            for category in categories
            if category
        )

        if category_path:
            category_counter[
                category_path
            ] += 1

    representative_category = (
        category_counter.most_common(1)[0][0]
        if category_counter
        else ""
    )

    return {
        "product_count": int(
            data.get("total") or 0
        ),
        "representative_category": (
            representative_category
        ),
        "category_sample_count": (
            category_counter.most_common(1)[0][1]
            if category_counter
            else 0
        ),
    }


def get_shopping_category_cached(
    keyword: str,
) -> dict[str, Any]:
    cache_key = normalize_keyword(
        keyword
    ).casefold()
    current_time = time.time()

    with _category_cache_lock:
        cached = _category_cache.get(
            cache_key
        )

        if (
            cached
            and current_time - cached[0]
            < CATEGORY_CACHE_SECONDS
        ):
            return {
                **cached[1],
                "category_cached": True,
            }

    result = fetch_shopping_category(
        keyword
    )

    with _category_cache_lock:
        _category_cache[cache_key] = (
            current_time,
            result,
        )

    return {
        **result,
        "category_cached": False,
    }


def analyze_keywords(
    keyword: str,
    related_limit: int = 20,
) -> dict[str, Any]:
    started_at = time.perf_counter()
    normalized_seed = normalize_keyword(
        keyword
    )

    statistics = fetch_keyword_statistics(
        normalized_seed
    )

    rows: list[dict[str, Any]] = []
    seen: set[str] = set()

    for item in statistics:
        related_keyword = normalize_keyword(
            item.get("relKeyword")
        )

        if not related_keyword:
            continue

        key = related_keyword.casefold()

        if key in seen:
            continue

        seen.add(key)

        pc_volume, pc_estimated = parse_volume(
            item.get("monthlyPcQcCnt")
        )
        mobile_volume, mobile_estimated = (
            parse_volume(
                item.get("monthlyMobileQcCnt")
            )
        )

        rows.append({
            "keyword": related_keyword,
            "pc_volume": pc_volume,
            "pc_volume_raw": str(
                item.get("monthlyPcQcCnt")
                or "0"
            ),
            "pc_estimated": pc_estimated,
            "mobile_volume": mobile_volume,
            "mobile_volume_raw": str(
                item.get("monthlyMobileQcCnt")
                or "0"
            ),
            "mobile_estimated": (
                mobile_estimated
            ),
            "total_volume": (
                pc_volume + mobile_volume
            ),
            "competition": str(
                item.get("compIdx") or ""
            ),
            "average_pc_clicks": float(
                item.get(
                    "monthlyAvePcClkCnt"
                ) or 0
            ),
            "average_mobile_clicks": float(
                item.get(
                    "monthlyAveMobileClkCnt"
                ) or 0
            ),
        })

    rows.sort(
        key=lambda row: (
            0
            if row["keyword"].casefold()
            == normalized_seed.casefold()
            else 1,
            -int(row["total_volume"]),
            row["keyword"],
        )
    )

    rows = rows[:related_limit]

    with ThreadPoolExecutor(
        max_workers=min(3, max(1, len(rows)))
    ) as executor:
        category_results = list(
            executor.map(
                lambda row: (
                    get_shopping_category_cached(
                        row["keyword"]
                    )
                ),
                rows,
            )
        )

    for row, category in zip(
        rows,
        category_results,
    ):
        row.update(category)

    seed_row = next(
        (
            row
            for row in rows
            if row["keyword"].casefold()
            == normalized_seed.casefold()
        ),
        rows[0] if rows else None,
    )

    return {
        "keyword": normalized_seed,
        "related_limit": related_limit,
        "count": len(rows),
        "elapsed_seconds": round(
            time.perf_counter() - started_at,
            3,
        ),
        "summary": (
            seed_row
            if seed_row
            else {
                "keyword": normalized_seed,
                "pc_volume": 0,
                "mobile_volume": 0,
                "total_volume": 0,
                "product_count": 0,
                "representative_category": "",
            }
        ),
        "keywords": rows,
    }



def safe_excel_text(value: Any) -> str:
    text = str(value or "").strip()

    if text.startswith(("=", "+", "-", "@")):
        return "'" + text

    return text


def display_export_volume(
    raw: Any,
    value: Any,
) -> int | str:
    normalized = str(raw or "").replace(
        " ",
        "",
    )

    if normalized in {"<10", "10미만"}:
        return "< 10"

    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def create_keyword_analysis_workbook(
    rows: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "키워드 분석"

    headers = [
        "키워드",
        "해시태그",
        "PC 검색량",
        "모바일 검색량",
        "총 검색량",
        "PC 평균 클릭",
        "모바일 평균 클릭",
        "쇼핑 상품 수",
        "경쟁도",
        "대표 카테고리",
        "카테고리 표본 수",
    ]

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0A84FF",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row in rows:
        keyword = safe_excel_text(
            row.get("keyword")
        )
        hashtag = (
            "#"
            + re.sub(
                r"\s+",
                "",
                keyword.lstrip("'"),
            )
            if keyword
            else ""
        )

        worksheet.append([
            keyword,
            safe_excel_text(hashtag),
            display_export_volume(
                row.get("pc_volume_raw"),
                row.get("pc_volume"),
            ),
            display_export_volume(
                row.get("mobile_volume_raw"),
                row.get("mobile_volume"),
            ),
            int(row.get("total_volume") or 0),
            float(
                row.get("average_pc_clicks")
                or 0
            ),
            float(
                row.get(
                    "average_mobile_clicks"
                )
                or 0
            ),
            int(row.get("product_count") or 0),
            safe_excel_text(
                row.get("competition")
            ),
            safe_excel_text(
                row.get(
                    "representative_category"
                )
            ),
            int(
                row.get(
                    "category_sample_count"
                )
                or 0
            ),
        ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )
    worksheet.row_dimensions[1].height = 24

    widths = [
        22,
        22,
        14,
        16,
        14,
        15,
        17,
        16,
        12,
        38,
        18,
    ]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        for column in ("E", "H", "K"):
            worksheet[
                f"{column}{row_number}"
            ].number_format = "#,##0"

        for column in ("F", "G"):
            worksheet[
                f"{column}{row_number}"
            ].number_format = "#,##0.0"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()
