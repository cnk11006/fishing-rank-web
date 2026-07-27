from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)


def safe_excel_value(value: Any) -> Any:
    """외부 문자열이 엑셀 수식으로 실행되지 않도록 처리합니다."""
    if value is None:
        return ""

    if isinstance(value, (int, float, bool)):
        return value

    if isinstance(value, list):
        value = " | ".join(str(item) for item in value)

    text = str(value)

    if text.startswith(("=", "+", "-", "@")):
        return "'" + text

    return text


def join_values(value: Any) -> str:
    if isinstance(value, list):
        return " | ".join(
            str(item).strip()
            for item in value
            if str(item).strip()
        )

    return str(value or "").strip()


def set_sheet_style(
    worksheet,
    freeze: str = "A2",
) -> None:
    worksheet.freeze_panes = freeze
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 28

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in worksheet.columns:
        letter = get_column_letter(
            column_cells[0].column
        )
        width = 10

        for cell in column_cells[:200]:
            value = str(cell.value or "")
            width = max(
                width,
                min(45, len(value) + 2),
            )

        worksheet.column_dimensions[letter].width = width


def add_summary_sheet(
    workbook: Workbook,
    title: str,
    rows: list[tuple[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(["항목", "값"])

    for label, value in rows:
        worksheet.append([
            safe_excel_value(label),
            safe_excel_value(value),
        ])

    set_sheet_style(worksheet)
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 55


def add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            safe_excel_value(value)
            for value in row
        ])

    set_sheet_style(worksheet)


def save_workbook(workbook: Workbook) -> bytes:
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def create_cross_purchase_workbook(
    payload: dict[str, Any],
) -> bytes:
    workbook = Workbook()
    summary = dict(payload.get("summary") or {})
    target_query = str(
        payload.get("target_query") or ""
    )

    add_summary_sheet(
        workbook,
        "분석 요약",
        [
            ("기준 검색어", target_query),
            (
                "업로드 파일 수",
                summary.get("uploaded_file_count", 0),
            ),
            (
                "정상 파일 수",
                summary.get("successful_file_count", 0),
            ),
            (
                "파일 오류 수",
                summary.get("file_error_count", 0),
            ),
            (
                "주문 원본 행 수",
                summary.get("order_row_count", 0),
            ),
            (
                "전체 주문 수",
                summary.get("total_order_count", 0),
            ),
            (
                "기준상품 구매 주문 수",
                summary.get("target_order_count", 0),
            ),
            (
                "연관상품 수",
                summary.get("result_count", 0),
            ),
            (
                "중복 제거 수",
                summary.get("duplicate_row_count", 0),
            ),
            (
                "취소·반품 제외 수",
                summary.get("excluded_status_count", 0),
            ),
            (
                "표시 상품 수 설정",
                summary.get("top_n", 0),
            ),
            (
                "최소 동시구매 주문 수",
                summary.get("min_orders", 0),
            ),
            (
                "처리 시간(초)",
                payload.get("elapsed_seconds", 0),
            ),
        ],
    )

    result_rows: list[list[Any]] = []

    for index, item in enumerate(
        payload.get("results") or [],
        start=1,
    ):
        result_rows.append([
            index,
            item.get("recommendation_grade", ""),
            item.get("recommendation_score", 0),
            item.get("product_id", ""),
            item.get("product_name", ""),
            item.get("together_order_count", 0),
            item.get("cross_purchase_rate", 0),
            item.get("support", 0),
            item.get("confidence", 0),
            item.get("lift", 0),
            item.get("overall_order_count", 0),
            item.get("together_quantity", 0),
            item.get("together_revenue", 0),
            "자사상품" if item.get("is_ours") else "",
            join_values(item.get("warnings")),
        ])

    add_table_sheet(
        workbook,
        "연관상품 결과",
        [
            "순위",
            "추천 등급",
            "추천 점수",
            "상품번호",
            "상품명",
            "동시구매 주문 수",
            "교차구매율(%)",
            "Support(%)",
            "Confidence(%)",
            "Lift",
            "전체 출현 주문 수",
            "동시구매 수량",
            "동시구매 매출",
            "자사 여부",
            "주의사항",
        ],
        result_rows,
    )

    error_rows = [
        [
            item.get("file_name", ""),
            item.get("message", ""),
        ]
        for item in payload.get("file_errors") or []
    ]

    add_table_sheet(
        workbook,
        "파일 오류",
        ["파일명", "오류 내용"],
        error_rows,
    )

    guide = dict(payload.get("analysis_guide") or {})

    add_summary_sheet(
        workbook,
        "지표 설명",
        [
            ("Support", guide.get("support", "")),
            ("Confidence", guide.get("confidence", "")),
            ("Lift", guide.get("lift", "")),
        ],
    )

    return save_workbook(workbook)


def create_candidate_workbook(
    payload: dict[str, Any],
) -> bytes:
    workbook = Workbook()
    summary = dict(payload.get("summary") or {})
    keywords = payload.get("keywords") or []

    add_summary_sheet(
        workbook,
        "분석 요약",
        [
            ("검색 키워드", join_values(keywords)),
            (
                "검색 키워드 수",
                summary.get("keyword_count", 0),
            ),
            (
                "마스터 상품 수",
                summary.get("master_product_count", 0),
            ),
            (
                "최종 후보 수",
                summary.get("candidate_count", 0),
            ),
            (
                "검색 오류 수",
                summary.get("error_count", 0),
            ),
            (
                "자사몰 제외 수",
                summary.get("excluded_our_store_count", 0),
            ),
            (
                "보유상품 제외 수",
                summary.get("excluded_owned_count", 0),
            ),
            (
                "상품번호 일치 제외 수",
                summary.get("excluded_product_id_count", 0),
            ),
            (
                "상품명 일치 제외 수",
                summary.get("excluded_exact_name_count", 0),
            ),
            (
                "유사상품 제외 수",
                summary.get("excluded_similar_count", 0),
            ),
            (
                "보유 가능성 검토 수",
                summary.get("ownership_review_count", 0),
            ),
            (
                "키워드별 수집 수",
                summary.get("max_results", 0),
            ),
            (
                "최종 표시 수 설정",
                summary.get("result_limit", 0),
            ),
            (
                "최소 검색량",
                summary.get("min_volume", 0),
            ),
            (
                "처리 시간(초)",
                payload.get("elapsed_seconds", 0),
            ),
        ],
    )

    result_rows: list[list[Any]] = []

    for index, item in enumerate(
        payload.get("results") or [],
        start=1,
    ):
        detail = dict(item.get("score_detail") or {})

        result_rows.append([
            index,
            item.get("recommendation_grade", ""),
            item.get("potential_score", 0),
            item.get("product_id", ""),
            item.get("product_name", ""),
            item.get("brand", ""),
            item.get("maker", ""),
            join_values(item.get("keywords")),
            item.get("volume_keyword", ""),
            item.get("search_volume", 0),
            item.get("best_rank", 0),
            item.get("representative_seller", ""),
            item.get("observed_seller_count", 0),
            item.get("representative_price", 0),
            item.get("lowest_price", 0),
            item.get("average_price", 0),
            item.get("highest_price", 0),
            item.get("category", ""),
            item.get("link", ""),
            item.get("ownership_confidence", 0),
            (
                "검토 필요"
                if item.get("ownership_review")
                else ""
            ),
            item.get("matched_owned_product", ""),
            item.get("ownership_match_reason", ""),
            detail.get("demand", 0),
            detail.get("exposure", 0),
            detail.get("relevance", 0),
            detail.get("category", 0),
            detail.get("seller_diversity", 0),
            detail.get("price_stability", 0),
            join_values(
                item.get("recommendation_reasons")
            ),
            join_values(item.get("warnings")),
        ])

    add_table_sheet(
        workbook,
        "사입후보 결과",
        [
            "순위",
            "추천 등급",
            "종합 점수",
            "상품번호",
            "상품명",
            "브랜드",
            "제조사",
            "발견 검색어",
            "검색량 기준 키워드",
            "월간 검색량",
            "쇼핑 최고 순위",
            "대표 판매처",
            "관측 판매처 수",
            "대표 가격",
            "최저가",
            "평균가",
            "최고가",
            "카테고리",
            "상품 링크",
            "보유 가능성(%)",
            "보유 검토",
            "비교 보유상품",
            "보유 판정 근거",
            "수요 점수",
            "노출 점수",
            "관련성 점수",
            "카테고리 점수",
            "판매처 다양성",
            "가격 안정성",
            "추천 사유",
            "주의사항",
        ],
        result_rows,
    )

    error_rows = [
        [
            item.get("keyword", ""),
            item.get("message", ""),
        ]
        for item in payload.get("errors") or []
    ]

    add_table_sheet(
        workbook,
        "검색 오류",
        ["검색어", "오류 내용"],
        error_rows,
    )

    return save_workbook(workbook)
